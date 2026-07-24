import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- styles ---------------------------------------------------------------
FONT = "Arial"
hdr_fill   = PatternFill("solid", fgColor="1A2B4A")
hdr_font   = Font(name=FONT, bold=True, color="FFFFFF", size=10)
title_font = Font(name=FONT, bold=True, color="1A2B4A", size=15)
sub_font   = Font(name=FONT, italic=True, color="444444", size=9)
blue_font  = Font(name=FONT, color="0000FF", size=10)      # editable input
black_font = Font(name=FONT, color="000000", size=10)      # formula
bold_font  = Font(name=FONT, bold=True, size=10)
fill_font  = Font(name=FONT, color="0000FF", size=10)
yellow     = PatternFill("solid", fgColor="FFF7CC")        # fill-in during procurement
tier_fill  = PatternFill("solid", fgColor="EAF0F8")
tot_fill   = PatternFill("solid", fgColor="D6E0EF")
thin  = Side(style="thin", color="B8C2D0")
med   = Side(style="medium", color="1A2B4A")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
topb   = Border(top=med)
wrap   = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")
money  = '$#,##0'

# ==========================================================================
# Sheet 1: README / Legend
# ==========================================================================
rm = wb.active
rm.title = "README"
rm.sheet_view.showGridLines = False
rm["A1"] = "UAV Harvest-Monitoring Deployment — Procurement Workbook"
rm["A1"].font = title_font
rm["A2"] = ("Bill of materials for an industrial test run of the RobustIDPS.ai / "
            "M1+M4+M6+M7 network-layer robustness models on a crop harvest-readiness UAV.")
rm["A2"].font = sub_font
rows = [
 ("", ""),
 ("HOW TO USE", ""),
 ("1", "Work on the 'BOM' sheet. Each row is one line item."),
 ("2", "Fill the YELLOW cells during procurement: Quoted Unit $, Vendor (confirmed), Lead time, Status."),
 ("3", "Adjust BLUE cells (Qty, Unit $ Low/High estimates) to your farm size / spec."),
 ("4", "BLACK cells are formulas (extended costs, totals) — do not overtype them."),
 ("5", "'Cost Summary' and 'OPEX' roll up automatically."),
 ("", ""),
 ("LEGEND", ""),
 ("Blue text", "Editable input — quantity and cost estimates you can change."),
 ("Yellow fill", "Fill in during procurement — quoted price, vendor, lead time, status."),
 ("Black text", "Formula — recalculates automatically."),
 ("", ""),
 ("PRIORITY", ""),
 ("Must", "Required for any test run."),
 ("Should", "Strongly recommended (assurance / EW validation)."),
 ("Optional", "Add per crop / field size / budget (thermal, LiDAR, extra sensors)."),
 ("", ""),
 ("EXAMPLE ROW (format reference — real rows are on the BOM sheet)", ""),
 ("Item", "Onboard AI chip — Jetson Orin NX 16GB module"),
 ("Qty", "1"),
 ("Unit $ Low / High", "500 / 700"),
 ("Quoted Unit $ (you fill)", "615"),
 ("Status (you fill)", "Ordered — PO#1042, ETA 3 wks"),
 ("", ""),
 ("NOTE ON PRICES", "All costs are indicative 2026 USD planning ranges, not quotes. "
                    "Confirm with vendors for your region. Cloud + GNSS simulator can be rented (see OPEX)."),
]
r = 4
for a,b in rows:
    rm[f"A{r}"] = a; rm[f"B{r}"] = b
    if a in ("HOW TO USE","LEGEND","PRIORITY","EXAMPLE ROW (format reference — real rows are on the BOM sheet)","NOTE ON PRICES"):
        rm[f"A{r}"].font = Font(name=FONT, bold=True, color="1A2B4A", size=11)
    else:
        rm[f"A{r}"].font = Font(name=FONT, bold=True, size=10)
        rm[f"B{r}"].font = Font(name=FONT, size=10)
    r += 1
rm["A10"].fill = PatternFill("solid", fgColor="EAF0F8")  # blue legend cell
rm["A10"].font = blue_font
rm["A11"].fill = yellow
rm.column_dimensions["A"].width = 30
rm.column_dimensions["B"].width = 85

# ==========================================================================
# Sheet 2: BOM
# ==========================================================================
bom = wb.create_sheet("BOM")
bom.sheet_view.showGridLines = False
headers = ["#","Tier / Group","Subsystem","Function","Representative part",
           "Example vendor","Qty","Unit $ Low","Unit $ High",
           "Ext $ Low","Ext $ High","Quoted Unit $","Ext Quoted $",
           "Lead (wks)","Priority","Status","Notes"]
# rows: (tier, subsystem, function, part, vendor, qty, low, high, lead, priority, notes)
data = [
 ("UAV Platform","Airframe","Multirotor, >=30 min, payload >=1.5 kg","DJI M350 RTK / Freefly Astro","DJI / Freefly",1,5000,12000,4,"Must","Fixed-wing/VTOL for >200 ha"),
 ("UAV Platform","Flight controller","Open autopilot (PX4/ArduPilot)","Holybro Pixhawk 6X / Cube Orange+","Holybro / CubePilot",1,250,450,3,"Must","May be integrated in airframe"),
 ("Onboard AI","Embedded compute","INT8 MambaShield + onboard NDVI","Jetson Orin NX 16GB module","NVIDIA / Seeed",1,500,700,6,"Must","~100 TOPS INT8, 10-25 W"),
 ("Onboard AI","Carrier board","Jetson carrier for UAV","ConnectTech Boson / Auvidea","ConnectTech",1,200,400,6,"Must","Pick low-SWaP carrier"),
 ("Onboard AI","Security co-accel (opt.)","Dedicated INT8 detector","Hailo-8 M.2 / Coral TPU","Hailo / Google",1,200,450,6,"Optional","2-3 W offload of M4"),
 ("Navigation","Anti-spoof GNSS","RTK + spoof/jam detection","Septentrio mosaic-X5 / u-blox ZED-F9P","Septentrio / u-blox",1,300,3000,6,"Must","Feeds C/N0+spoof flags to M4/M6"),
 ("Navigation","GNSS antenna","Multiband / anti-jam","CRPA or multiband helical","Tallysman / Septentrio",1,200,2500,4,"Should","CRPA for serious EW"),
 ("Datalink","C2 MANET radio","Encrypted mesh C2 (monitored link)","Silvus StreamCaster / Doodle Labs","Silvus / Doodle Labs",1,3000,9000,8,"Must","The 'network layer' M1/M4 inspect"),
 ("Datalink","BVLOS backhaul","4G/5G modem + NTRIP","Sierra/Quectel 5G modem","Sierra Wireless",1,150,400,3,"Should","Telemetry + RTK corrections"),
 ("Datalink","Backup telemetry","Fail-safe C2","RFD900x","RFDesign",1,250,500,3,"Should","Redundant link"),
 ("Payload","Multispectral camera","NDVI/NDRE harvest-due sensing","MicaSense Altum-PT / RedEdge-P","AgEagle/MicaSense",1,8000,13000,8,"Must","Core harvest-readiness sensor"),
 ("Payload","Radiometric calibration","Downwelling sensor + panel","MicaSense DLS 2 + reflectance panel","AgEagle/MicaSense",1,500,1500,8,"Must","Essential for valid NDVI"),
 ("Payload","RGB camera (opt.)","High-res scouting / ground truth","Sony a7R / integrated","Sony",1,1000,6000,4,"Optional",""),
 ("Payload","Thermal (opt.)","Crop water stress","FLIR Vue TZ20","Teledyne FLIR",1,3000,8000,6,"Optional",""),
 ("Payload","LiDAR (opt.)","Biomass / canopy height","DJI Zenmuse L2 / Livox","DJI / Livox",1,3000,12000,6,"Optional",""),
 ("Payload","Gimbal","Payload stabilisation","Gremsy / integrated","Gremsy",1,300,2500,4,"Should",""),
 ("Edge Gateway","Tier-2 compute","Full RobustIDPS.ai + M1/M6/M7","Jetson AGX Orin 64GB (rugged)","NVIDIA / ConnectTech",1,2000,5000,6,"Must","~275 TOPS; full crop-maturity model"),
 ("Ground","RTK base station","cm-level corrections","Emlid Reach RS3","Emlid",1,2000,3500,3,"Should","Or NTRIP subscription (OPEX)"),
 ("Ground","Ground control station","Mission planning + monitoring","Rugged laptop/tablet + QGC","Getac / Panasonic",1,1500,4000,3,"Must","QGroundControl / Mission Planner"),
 ("Ground","Field networking","Router + PoE switch + UPS","Teltonika RUTX + PoE switch","Teltonika",1,400,900,3,"Should",""),
 ("Ground","Field power","Portable power + solar/generator","2 kWh power station + solar","EcoFlow / Bluetti",1,800,2500,3,"Should",""),
 ("Security","Secure element","Onboard key store / signed updates","Microchip ATECC608 (board)","Microchip",1,20,60,3,"Must","Verifies PQC-signed models"),
 ("Security","HSM","Sign Kyber-1024 / Dilithium updates","YubiHSM 2","Yubico",1,650,950,3,"Should","PQC update channel at gateway"),
 ("Test/Validation","Spectrum analyzer","Measure real J/S + noise floor","Signal Hound BB60 / tinySA Ultra","Signal Hound",1,500,3500,3,"Should","Validate field vs benchmark"),
 ("Test/Validation","GNSS sim (rent)","Cabled spoof/jam replay","LabSat 3 / Spirent","Racelogic / Spirent",1,1000,20000,4,"Should","Rental OK; range/Faraday only"),
 ("Test/Validation","SDR","RF characterisation","USRP B210 / HackRF","Ettus / Great Scott",1,300,3000,3,"Optional",""),
 ("Consumables","Flight batteries","Smart LiPo/Li-ion packs","OEM smart batteries","OEM",8,200,600,3,"Must","Rotate 6-10 packs"),
 ("Consumables","Charging","Multi-charger + LiPo bags","Fast multi-charger","OEM",1,300,1200,3,"Must",""),
 ("Tools","Assembly toolkit","Drivers, solder, multimeter, crimpers","Precision toolkit","iFixit / Weller",1,300,1500,2,"Must",""),
 ("Logistics","Transit cases","Rugged transport","Pelican / SKB","Pelican",3,150,400,3,"Should",""),
 ("Logistics","Field accessories","Landing pad, GCP targets, ND filters","Assorted","Assorted",1,200,800,2,"Should",""),
 ("Safety","Safety kit","Li-ion extinguisher, first aid, hi-vis","Safety kit","Assorted",1,200,600,2,"Must",""),
]

# header row
hr = 1
for c,h in enumerate(headers,1):
    cell = bom.cell(hr,c,h); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = border
first = 2
for i,row in enumerate(data):
    r = first+i
    tier,sub,func,part,vend,qty,low,high,lead,pri,notes = row
    vals = [i+1,tier,sub,func,part,vend,qty,low,high,
            f"=G{r}*H{r}", f"=G{r}*I{r}", None, f"=G{r}*L{r}",
            lead,pri,None,notes]
    for c,v in enumerate(vals,1):
        cell = bom.cell(r,c,v); cell.border = border; cell.font = black_font
        cell.alignment = wrap if c in (3,4,5,17) else Alignment(vertical="top")
    # blue inputs: qty(7), low(8), high(9)
    for c in (7,8,9): bom.cell(r,c).font = blue_font
    # yellow fill-ins: quoted unit(12), status(16); lead(14) editable
    for c in (12,16):
        bom.cell(r,c).fill = yellow; bom.cell(r,c).font = fill_font
    bom.cell(r,14).font = blue_font
    # money format
    for c in (8,9,10,11,12,13): bom.cell(r,c).number_format = money
    bom.cell(r,7).alignment = center
    bom.cell(r,15).alignment = center

last = first+len(data)-1
tr = last+1
bom.cell(tr,1,"TOTAL (est. range)").font = bold_font
bom.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=9)
bom.cell(tr,1).alignment = Alignment(horizontal="right")
for c in (10,11,13):
    L = get_column_letter(c)
    cell = bom.cell(tr,c,f"=SUM({L}{first}:{L}{last})")
    cell.font = bold_font; cell.number_format = money; cell.fill = tot_fill
    cell.border = Border(top=med,bottom=med)
for c in range(1,18):
    bom.cell(tr,c).fill = tot_fill

widths = [4,14,16,26,30,20,5,10,10,11,11,12,12,8,9,20,34]
for c,w in enumerate(widths,1):
    bom.column_dimensions[get_column_letter(c)].width = w
bom.freeze_panes = "A2"
bom.auto_filter.ref = f"A1:Q{last}"

# ==========================================================================
# Sheet 3: Cost Summary
# ==========================================================================
cs = wb.create_sheet("Cost Summary")
cs.sheet_view.showGridLines = False
cs["A1"] = "Cost Summary"; cs["A1"].font = title_font
cs["A2"] = "Rolls up from the BOM sheet. Estimated ranges; Quoted column fills as you procure."
cs["A2"].font = sub_font

tiers = ["UAV Platform","Onboard AI","Navigation","Datalink","Payload","Edge Gateway",
         "Ground","Security","Test/Validation","Consumables","Tools","Logistics","Safety"]
sh = 4
for c,h in enumerate(["Tier / Group","Est $ Low","Est $ High","Quoted $"],1):
    cell = cs.cell(sh,c,h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border=border
    cell.alignment = center
for i,t in enumerate(tiers):
    r = sh+1+i
    cs.cell(r,1,t).font = black_font; cs.cell(r,1).border=border
    cs.cell(r,2,f'=SUMIFS(BOM!$J$2:$J$33,BOM!$B$2:$B$33,$A{r})').number_format=money
    cs.cell(r,3,f'=SUMIFS(BOM!$K$2:$K$33,BOM!$B$2:$B$33,$A{r})').number_format=money
    cs.cell(r,4,f'=SUMIFS(BOM!$M$2:$M$33,BOM!$B$2:$B$33,$A{r})').number_format=money
    for c in (2,3,4):
        cs.cell(r,c).font=black_font; cs.cell(r,c).border=border
last_t = sh+len(tiers)
tr2 = last_t+1
cs.cell(tr2,1,"CAPEX TOTAL").font=bold_font
for c in (2,3,4):
    L=get_column_letter(c)
    cs.cell(tr2,c,f"=SUM({L}{sh+1}:{L}{last_t})").number_format=money
    cs.cell(tr2,c).font=bold_font; cs.cell(tr2,c).fill=tot_fill
    cs.cell(tr2,c).border=Border(top=med,bottom=med)
cs.cell(tr2,1).fill=tot_fill; cs.cell(tr2,1).border=Border(top=med,bottom=med)

# by priority
ph = tr2+3
cs.cell(ph,1,"By priority").font=Font(name=FONT,bold=True,size=11,color="1A2B4A")
for c,h in enumerate(["Priority","Est $ Low","Est $ High","Quoted $"],1):
    cell=cs.cell(ph+1,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=border; cell.alignment=center
for i,p in enumerate(["Must","Should","Optional"]):
    r=ph+2+i
    cs.cell(r,1,p).font=black_font; cs.cell(r,1).border=border
    cs.cell(r,2,f'=SUMIFS(BOM!$J$2:$J$33,BOM!$O$2:$O$33,$A{r})').number_format=money
    cs.cell(r,3,f'=SUMIFS(BOM!$K$2:$K$33,BOM!$O$2:$O$33,$A{r})').number_format=money
    cs.cell(r,4,f'=SUMIFS(BOM!$M$2:$M$33,BOM!$O$2:$O$33,$A{r})').number_format=money
    for c in (2,3,4): cs.cell(r,c).font=black_font; cs.cell(r,c).border=border
# lean build note
cs.cell(ph+6,1,'Lean first run = "Must" only. Full assurance = Must + Should. Optional = crop/field-specific add-ons.').font=sub_font

for c,w in zip("ABCD",[20,14,14,14]):
    cs.column_dimensions[c].width=w

# ==========================================================================
# Sheet 4: OPEX (recurring)
# ==========================================================================
op = wb.create_sheet("OPEX")
op.sheet_view.showGridLines = False
op["A1"]="Recurring Costs (OPEX)"; op["A1"].font=title_font
op["A2"]="Monthly services and consumable-rate items. Fill blue estimates for your setup."
op["A2"].font=sub_font
for c,h in enumerate(["Item","Purpose","$ / month (Low)","$ / month (High)","$ / year (Low)","$ / year (High)"],1):
    cell=op.cell(4,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=border; cell.alignment=center
opex=[
 ("Cloud GPU (FedGTD + PQC signing)","Tier-3 aggregation & model signing",300,1500),
 ("NTRIP correction subscription","RTK corrections (if no own base)",20,100),
 ("Connectivity SIM / data plan","BVLOS 4G/5G backhaul",30,150),
 ("Liability / hull insurance","Ops insurance",100,400),
 ("Software maintenance / updates","Model retraining, container upkeep",100,500),
 ("GNSS simulator rental (test window)","EW validation, per month when testing",0,3000),
]
fr=5
for i,(it,pu,lo,hi) in enumerate(opex):
    r=fr+i
    op.cell(r,1,it).font=black_font; op.cell(r,2,pu).font=black_font
    op.cell(r,3,lo).font=blue_font; op.cell(r,4,hi).font=blue_font
    op.cell(r,5,f"=C{r}*12").font=black_font; op.cell(r,6,f"=D{r}*12").font=black_font
    for c in range(1,7):
        op.cell(r,c).border=border
        if c>=3: op.cell(r,c).number_format=money
lr=fr+len(opex)-1; trr=lr+1
op.cell(trr,1,"TOTAL OPEX").font=bold_font
op.merge_cells(start_row=trr,start_column=1,end_row=trr,end_column=2)
op.cell(trr,1).alignment=Alignment(horizontal="right")
for c in (3,4,5,6):
    L=get_column_letter(c)
    op.cell(trr,c,f"=SUM({L}{fr}:{L}{lr})").number_format=money
    op.cell(trr,c).font=bold_font; op.cell(trr,c).fill=tot_fill; op.cell(trr,c).border=Border(top=med,bottom=med)
op.cell(trr,1).fill=tot_fill; op.cell(trr,2).fill=tot_fill
for c,w in zip("ABCDEF",[34,34,16,16,14,14]):
    op.column_dimensions[c].width=w

wb.save("UAV_Deployment_Procurement.xlsx")
print("saved")
